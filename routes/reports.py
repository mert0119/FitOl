import io
from flask import Blueprint, render_template, request, send_file
from flask_login import login_required, current_user
from datetime import date, timedelta
from sqlalchemy import func
from models import db
from models.food_log import FoodLog
from models.exercise_log import ExerciseLog
from models.water_log import WaterLog
from models.measurement import Measurement

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports')
@login_required
def index():
    period = request.args.get('period', 'weekly')
    today = date.today()

    if period == 'daily':
        start_date = today
        title = f"Günlük Rapor — {today.strftime('%d/%m/%Y')}"
    elif period == 'monthly':
        start_date = today - timedelta(days=30)
        title = f"Aylık Rapor — Son 30 Gün"
    else:
        start_date = today - timedelta(days=7)
        title = f"Haftalık Rapor — Son 7 Gün"
        period = 'weekly'

    days_count = (today - start_date).days + 1

    # Kalori istatistikleri
    calorie_stats = db.session.query(
        func.coalesce(func.sum(FoodLog.calories), 0),
        func.coalesce(func.avg(FoodLog.calories), 0),
        func.coalesce(func.sum(FoodLog.protein), 0),
        func.coalesce(func.sum(FoodLog.carbs), 0),
        func.coalesce(func.sum(FoodLog.fat), 0)
    ).filter(
        FoodLog.user_id == current_user.id,
        FoodLog.date >= start_date,
        FoodLog.date <= today
    ).first()

    # Egzersiz istatistikleri
    exercise_stats = db.session.query(
        func.coalesce(func.sum(ExerciseLog.calories_burned), 0),
        func.coalesce(func.sum(ExerciseLog.duration), 0),
        func.count(ExerciseLog.id)
    ).filter(
        ExerciseLog.user_id == current_user.id,
        ExerciseLog.date >= start_date,
        ExerciseLog.date <= today
    ).first()

    # Su istatistikleri
    water_stats = db.session.query(
        func.coalesce(func.sum(WaterLog.amount_ml), 0)
    ).filter(
        WaterLog.user_id == current_user.id,
        WaterLog.date >= start_date,
        WaterLog.date <= today
    ).scalar()

    # Günlük kalori grafik verisi
    daily_data = {'labels': [], 'calories_in': [], 'calories_out': [], 'water': []}
    for i in range(days_count):
        d = start_date + timedelta(days=i)
        daily_data['labels'].append(d.strftime('%d/%m'))

        cal_in = db.session.query(
            func.coalesce(func.sum(FoodLog.calories), 0)
        ).filter(FoodLog.user_id == current_user.id, FoodLog.date == d).scalar()
        daily_data['calories_in'].append(round(cal_in))

        cal_out = db.session.query(
            func.coalesce(func.sum(ExerciseLog.calories_burned), 0)
        ).filter(ExerciseLog.user_id == current_user.id, ExerciseLog.date == d).scalar()
        daily_data['calories_out'].append(round(cal_out))

        water = db.session.query(
            func.coalesce(func.sum(WaterLog.amount_ml), 0)
        ).filter(WaterLog.user_id == current_user.id, WaterLog.date == d).scalar()
        daily_data['water'].append(round(water))

    # En çok yenen yemekler
    top_foods = db.session.query(
        FoodLog.food_name,
        func.count(FoodLog.id).label('count'),
        func.avg(FoodLog.calories).label('avg_cal')
    ).filter(
        FoodLog.user_id == current_user.id,
        FoodLog.date >= start_date
    ).group_by(FoodLog.food_name).order_by(func.count(FoodLog.id).desc()).limit(5).all()

    # En çok yapılan egzersizler
    top_exercises = db.session.query(
        ExerciseLog.exercise_name,
        func.count(ExerciseLog.id).label('count'),
        func.sum(ExerciseLog.calories_burned).label('total_cal')
    ).filter(
        ExerciseLog.user_id == current_user.id,
        ExerciseLog.date >= start_date
    ).group_by(ExerciseLog.exercise_name).order_by(func.count(ExerciseLog.id).desc()).limit(5).all()

    return render_template('reports.html',
                           period=period,
                           title=title,
                           days_count=days_count,
                           calorie_stats={
                               'total': round(calorie_stats[0]),
                               'avg_daily': round(calorie_stats[0] / max(days_count, 1)),
                               'total_protein': round(calorie_stats[2]),
                               'total_carbs': round(calorie_stats[3]),
                               'total_fat': round(calorie_stats[4])
                           },
                           exercise_stats={
                               'total_calories': round(exercise_stats[0]),
                               'total_duration': round(exercise_stats[1]),
                               'total_sessions': exercise_stats[2]
                           },
                           water_stats={
                               'total_ml': round(water_stats),
                               'avg_daily_ml': round(water_stats / max(days_count, 1)),
                               'avg_daily_litre': round(water_stats / max(days_count, 1) / 1000, 1)
                           },
                           daily_data=daily_data,
                           top_foods=top_foods,
                           top_exercises=top_exercises)


@reports_bp.route('/reports/export/excel')
@login_required
def export_excel():
    period = request.args.get('period', 'weekly')
    today = date.today()

    if period == 'daily':
        start_date = today
    elif period == 'monthly':
        start_date = today - timedelta(days=30)
    else:
        start_date = today - timedelta(days=7)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Stiller
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)

    # Sayfa 1: Yemek Logları
    ws_food = wb.active
    ws_food.title = "Yemek Logları"
    food_headers = ["Tarih", "Öğün", "Yemek", "Kalori", "Protein (g)", "Karb (g)", "Yağ (g)", "Porsiyon"]
    style_header(ws_food, food_headers)

    food_logs = FoodLog.query.filter(
        FoodLog.user_id == current_user.id,
        FoodLog.date >= start_date,
        FoodLog.date <= today
    ).order_by(FoodLog.date.desc()).all()

    for i, log in enumerate(food_logs, 2):
        ws_food.cell(row=i, column=1, value=log.date.strftime('%d/%m/%Y')).border = thin_border
        ws_food.cell(row=i, column=2, value=log.meal_type or '').border = thin_border
        ws_food.cell(row=i, column=3, value=log.food_name).border = thin_border
        ws_food.cell(row=i, column=4, value=round(log.calories or 0)).border = thin_border
        ws_food.cell(row=i, column=5, value=round(log.protein or 0, 1)).border = thin_border
        ws_food.cell(row=i, column=6, value=round(log.carbs or 0, 1)).border = thin_border
        ws_food.cell(row=i, column=7, value=round(log.fat or 0, 1)).border = thin_border
        ws_food.cell(row=i, column=8, value=log.portion or 1).border = thin_border

    # Sayfa 2: Egzersiz Logları
    ws_ex = wb.create_sheet("Egzersiz Logları")
    ex_headers = ["Tarih", "Egzersiz", "Kategori", "Süre (dk)", "Set", "Tekrar", "Ağırlık (kg)", "Kalori"]
    style_header(ws_ex, ex_headers)

    ex_logs = ExerciseLog.query.filter(
        ExerciseLog.user_id == current_user.id,
        ExerciseLog.date >= start_date,
        ExerciseLog.date <= today
    ).order_by(ExerciseLog.date.desc()).all()

    for i, log in enumerate(ex_logs, 2):
        ws_ex.cell(row=i, column=1, value=log.date.strftime('%d/%m/%Y')).border = thin_border
        ws_ex.cell(row=i, column=2, value=log.exercise_name).border = thin_border
        ws_ex.cell(row=i, column=3, value=log.category or '').border = thin_border
        ws_ex.cell(row=i, column=4, value=log.duration or 0).border = thin_border
        ws_ex.cell(row=i, column=5, value=log.sets or 0).border = thin_border
        ws_ex.cell(row=i, column=6, value=log.reps or 0).border = thin_border
        ws_ex.cell(row=i, column=7, value=log.weight_kg or 0).border = thin_border
        ws_ex.cell(row=i, column=8, value=round(log.calories_burned or 0)).border = thin_border

    # Sayfa 3: Su Logları
    ws_water = wb.create_sheet("Su Logları")
    water_headers = ["Tarih", "Miktar (ml)"]
    style_header(ws_water, water_headers)

    water_logs = WaterLog.query.filter(
        WaterLog.user_id == current_user.id,
        WaterLog.date >= start_date,
        WaterLog.date <= today
    ).order_by(WaterLog.date.desc()).all()

    for i, log in enumerate(water_logs, 2):
        ws_water.cell(row=i, column=1, value=log.date.strftime('%d/%m/%Y')).border = thin_border
        ws_water.cell(row=i, column=2, value=log.amount_ml or 0).border = thin_border

    # Dosyayı belleğe yaz
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"FitOl_Rapor_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
